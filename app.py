import os
os.environ['NO_PROXY'] = '*'

import streamlit as st
import pdfplumber
import pandas as pd
import json
import io
import zipfile
import re
import time
import uuid
import base64
from dotenv import load_dotenv
from rembg import remove
from PIL import Image
from openai import OpenAI
import httpx
from google.cloud import storage
from google.oauth2 import service_account

# --- 1. CONFIGURATION & AUTHENTICATION ---
load_dotenv()

def get_config(key, default=None):
    try:
        return st.secrets[key]
    except Exception:
        return os.getenv(key, default)

# Pull keys from Streamlit Secrets (Cloud) or fallback to local .env
OPENAI_API_KEY = get_config("OPENAI_API_KEY")
GCP_PROJECT_ID = get_config("GCP_PROJECT_ID", "caramel-goal-473111-t3")
GCS_BUCKET_NAME = get_config("GCS_BUCKET_NAME", "gyaanbuddy-media")

# Initialize OpenAI Client (Bypassing Windows/Cloud Proxies)
if OPENAI_API_KEY:
    try:
        custom_http = httpx.Client(proxy=None, timeout=60.0)
        openai_client = OpenAI(
            api_key=OPENAI_API_KEY, 
            http_client=custom_http,
            max_retries=3
        )
    except Exception as e:
        openai_client = None
        print(f"OpenAI Initialization Error: {e}")
else:
    openai_client = None

# Initialize GCS Client securely (supports Streamlit Cloud Secrets or local file)
try:
    has_gcp_secret = False
    try:
        has_gcp_secret = "gcp_service_account" in st.secrets
    except Exception:
        pass

    if has_gcp_secret:
        # For Streamlit Community Cloud Deployment
        creds_dict = dict(st.secrets["gcp_service_account"])
        credentials = service_account.Credentials.from_service_account_info(creds_dict)
        storage_client = storage.Client(project=GCP_PROJECT_ID, credentials=credentials)
    else:
        # For Local Development (using gcp_credentials.json file)
        storage_client = storage.Client(project=GCP_PROJECT_ID)
except Exception as e:
    storage_client = None
    print(f"GCS Auth Error: {e}")

# Initialize GCS Client for Cloud Storage Uploads
try:
    storage_client = storage.Client(project=GCP_PROJECT_ID)
except Exception as e:
    storage_client = None
    print(f"GCS Auth Error: {e}")

# --- 2. TEXT EXTRACTION FUNCTIONS ---
def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]

def extract_pdf_streams(uploaded_files):
    pdf_streams = []
    for file in uploaded_files:
        filename = file.name.lower()
        if filename.endswith(".zip"):
            try:
                with zipfile.ZipFile(file) as zf:
                    valid_pdf_names = [
                        name for name in zf.namelist() 
                        if name.lower().endswith(".pdf") 
                        and not name.startswith("__MACOSX/") 
                        and not name.split("/")[-1].startswith("._")
                    ]
                    valid_pdf_names.sort(key=natural_sort_key)
                    for name in valid_pdf_names:
                        pdf_bytes = io.BytesIO(zf.read(name))
                        pdf_bytes.name = name.split("/")[-1]
                        pdf_streams.append(pdf_bytes)
            except Exception as e:
                st.error(f"Error reading ZIP file {file.name}: {e}")
        elif filename.endswith(".pdf"):
            pdf_streams.append(file)
            
    pdf_streams.sort(key=lambda x: natural_sort_key(x.name))
    return pdf_streams

def extract_text_from_pdf(pdf_file_obj):
    full_text = []
    with pdfplumber.open(pdf_file_obj) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                full_text.append(f"--- PAGE {i+1} ---\n" + text)
    return "\n\n".join(full_text)

def process_text_with_ai(raw_text, filename, max_retries=3):
    """Passes chapter text AND filename to GPT-4o-mini to extract structured data."""
    if not openai_client:
        st.error(f"🚨 OpenAI Client is not initialized. Please check your API key.")
        return []
        
    prompt = f"""
    You are an expert curriculum and textbook indexing system. I am providing you with the text of a school textbook chapter.
    
    Source File Name: {filename}
    
    Your task is to comprehensively extract the Chapter Number, Chapter Name, and ALL OF ITS SUBTOPICS.
    
    CRITICAL INSTRUCTIONS:
    1. Extract the exact Chapter Number and Chapter Title. 
       - **CRUCIAL RULE:** Use the 'Source File Name' provided above as the ultimate source of truth for the Chapter Number (e.g., if the file is named 'ch11_something.pdf', the Chapter Number MUST be "11"). Do not let stray page numbers or typos in the text confuse you.
    2. Extract all main subtopics and section headings.
    3. Generate sequential Subtopic IDs starting with the exact Chapter Number (e.g., 11.1, 11.2, 11.3).
    4. Ignore questions, exercises, 'Let's Check', 'Activities', 'Did You Know' sidebars, and generic page filler.
    
    Output STRICTLY a valid JSON array of objects. Do not include markdown formatting like ```json.
    Format each item exactly like this:
    [
      {{
        "Chapter Number": "1",
        "Chapter Name": "Locating Places and Reading Maps",
        "Subtopic ID": "1.1",
        "Subtopic Name": "Shape of Earth"
      }}
    ]
    
    Textbook Content:
    {raw_text}
    """

    for attempt in range(1, max_retries + 1):
        try:
            response = openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a precise data extraction assistant. Always output clean, raw JSON."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1
            )
            
            content = response.choices[0].message.content.strip()
            if content.startswith("```json"):
                content = content.replace("```json", "", 1)
            if content.endswith("```"):
                content = content[:-3]
                
            return json.loads(content.strip())
            
        except Exception as e:
            if attempt < max_retries: 
                time.sleep(3 * attempt)
            else: 
                st.error(f"🚨 Text Extraction API Error on {filename}: {repr(e)}")
                return []

# --- 3. ICON GENERATION FUNCTIONS ---
SYSTEM_DESIGN_PROMPT = """
You are a **world-class 3D icon designer** creating icons for an educational platform (classes 4–12).

Your job is to generate **minimal, high-quality 3D icons** that act as **visual learning anchors**.

---
## 🎯 OBJECTIVE
For every input (Subject + Chapter), you must:
* Convert the chapter into **one clear visual concept**
* Design a **simple, recognizable 3D icon**
* Maintain **strict color consistency**
* Ensure icons are **visually lively (not monotone)**
* Remove background **there should be no background**

---
## 🎨 SUBJECT COLOR SYSTEM (PRIMARY – STRICT)
Use ONLY the assigned base color:
* Maths → #1A3FC4
* Science → #7B5EA7
* English → #22C55E
* Geography → #2E7D32
* Economics → #FF8C00
* Civics → #C62828
* History → #800000
* Computer / IT → #1FB7EB
* Commerce → #3D2DB5
* EVS → #6B8E23
* Psychology → #FF6B35
❗ Do NOT mix subject colors
❗ Base color must dominate (80–85%)

---
## 🎨 ACCENT COLOR SYSTEM (MANDATORY)
Each subject MUST use its predefined accent color (10–15%):
* Maths → #FFD54F (Warm Amber)
* Science → #4FC3F7 (Sky Cyan)
* English → #FFC107 (Golden Yellow)
* Geography → #FDD835 (Sun Yellow)
* Economics → #42A5F5 (Cool Blue)
* Civics → #E0E0E0 (Soft Silver)
* History → #D4AF37 (Antique Gold)
* Computer → #7E57C2 (Soft Violet)
* Commerce → #26C6DA (Aqua Cyan)
* EVS → #81D4FA (Light Sky)
* Psychology → #BA68C8 (Soft Purple)

### Accent Rules:
* Use only ONE accent color
* Use on small details only (edges, highlights, secondary elements)
* Must NOT overpower base color

---
## 🎨 BRAND CONSISTENCY (SUBTLE – 5%)
Add a very subtle tint using:
* #00167A OR
* #1B00AD

Use as:
* soft shadow tint OR
* slight edge highlight
❗ Must be barely noticeable
❗ Must NOT compete with main colors

---
## 🧱 DESIGN STYLE
* ADAPT THE STYLE based on the provided Target Audience (Playful/Chunky for younger classes, Sleek/Mature for older classes).
* Rounded edges
* Smooth matte finish
* Minimal (1–2 objects only)

---
## 💡 LIGHTING
* Soft light from top-left
* Gentle shadow below
* No harsh contrast

---
## 📐 COMPOSITION
* Centered
* Less spacing
* Fill the frame
* No clutter
* No background elements

---
## 🎨 BACKGROUND
* Light neutral (#E8EDF8 or off-white)

---
## 🧠 CONCEPT RULE
Before generating, decide:
“What is the ONE visual that represents this chapter?”

Then:
* Use symbols, not scenes
* Avoid storytelling
* Keep it instantly understandable

---
## ✍️ TEXT RULE
* Avoid text inside icons
* Only if absolutely necessary (max 1–2 words)

---
## ❌ DO NOT
* Use multiple subject colors
* Add unnecessary elements
* Create complex scenes
* Use flat design
* Break consistency

---
## 📦 OUTPUT FORMAT
Concept:
[One-line explanation]

Icon Prompt:
"A clean 3D icon of [main object], representing [concept], primarily in [base color hex] with subtle [accent color hex] highlights, [Insert exact adaptive styling details here based on the target audience], centered composition, fill frame, light neutral background, modern educational app style"
"""

def generate_icon_prompt(subject: str, chapter: str, filename: str):
    """Uses GPT-4o-mini as a design agent to craft the perfect image prompt based on class level."""
    if not openai_client:
        raise ValueError("OpenAI client not initialized.")
        
    class_match = re.search(r'(?:class|grade|std)[\s_-]*(\d+)', filename, re.IGNORECASE)
    
    if class_match:
        class_num = int(class_match.group(1))
        if class_num <= 7:
            audience_style = "Target Audience: Young students (Class 4-7). Style Rules: Make it a cute miniature diorama. Use thick, chunky, toy-like exaggerated proportions. Materials should resemble soft clay or vibrant plastic. Keep it playful and highly approachable."
        else:
            audience_style = "Target Audience: Older students (Class 8-12). Style Rules: Make it mature, sophisticated, and sleek. Use precision geometric forms, frosted glass, and premium matte materials inspired by high-end macOS/iOS app icons. Avoid making it look like a toy."
    else:
        audience_style = "Target Audience: General students. Style Rules: Balanced modern 3D geometry, premium and polished."

    response = openai_client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": SYSTEM_DESIGN_PROMPT},
            {"role": "user", "content": f"Subject: {subject}\nChapter: {chapter}\n{audience_style}"}
        ],
        temperature=0.3
    )
    
    output_text = response.choices[0].message.content.strip()
    prompt_match = re.search(r"Icon Prompt:\s*(.*)", output_text, re.DOTALL | re.IGNORECASE)
    return prompt_match.group(1).strip().strip('"').strip("'") if prompt_match else output_text.strip()

def render_openai_image(prompt_text: str) -> Image.Image:
    """Renders the asset natively using OpenAI's latest gpt-image-2 model."""
    if not openai_client:
        raise ValueError("OpenAI client not initialized. Check OPENAI_API_KEY.")
        
    response = openai_client.images.generate(
        model="gpt-image-2",
        prompt=prompt_text,
        n=1,
        size="1024x1024"
    )
    
    image_bytes = base64.b64decode(response.data[0].b64_json)
    return Image.open(io.BytesIO(image_bytes))

# --- 4. CLOUD STORAGE UPLOAD ---
def upload_to_gcs(image_bytes: bytes, bucket_name: str, destination_blob_name: str) -> str:
    if not storage_client:
        return "GCS_NOT_CONFIGURED_URL"
    bucket = storage_client.bucket(bucket_name)
    blob = bucket.blob(destination_blob_name)
    blob.upload_from_string(image_bytes, content_type='image/png')
    return f"https://storage.googleapis.com/{bucket_name}/{destination_blob_name}"

# --- 5. STREAMLIT UI & MAIN PIPELINE ---
# --- 5. STREAMLIT UI & MAIN PIPELINE ---
st.set_page_config(page_title="Universal Asset & Index Generator", layout="wide")
st.title("📚 Unified Textbook Indexer & Icon Studio")
st.markdown("Automated text extraction mapped directly to OpenAI visual anchors.")

subject_input = st.text_input("Subject (For Icon Colors)", value="", placeholder="e.g. Science, History...")

uploaded_files = st.file_uploader("Upload PDF or ZIP files", type=["pdf", "zip"], accept_multiple_files=True)

if uploaded_files and st.button("Extract Data & Generate Master Excel", type="primary"):
    discovered_pdfs = extract_pdf_streams(uploaded_files)
    master_data = []
    
    progress_bar = st.progress(0, text="Starting pipeline...")
    
    for idx, pdf_file in enumerate(discovered_pdfs):
        progress_bar.progress((idx + 1) / len(discovered_pdfs), text=f"Processing `{pdf_file.name}`...")
        
        # Step 1: Text Extraction
        raw_text = extract_text_from_pdf(pdf_file)
        
        if len(raw_text.strip()) < 50:
            st.warning(f"⚠️ `{pdf_file.name}` has no selectable text (it might be a scanned image). Skipping.")
            continue
            
        chapter_data = process_text_with_ai(raw_text, pdf_file.name)
        
        if not chapter_data:
            st.warning(f"⚠️ No structural data could be extracted from `{pdf_file.name}`. Skipping to next file.")
            continue
            
        # Step 2: Icon Generation
        chapter_name = chapter_data[0].get("Chapter Name", "Unknown Chapter")
        st.toast(f"Designing icon for: {chapter_name}")
        
        try:
            icon_prompt = generate_icon_prompt(subject_input, chapter_name, pdf_file.name)
            raw_image = render_openai_image(icon_prompt)
            transparent_icon = remove(raw_image)
            
            img_byte_arr = io.BytesIO()
            transparent_icon.save(img_byte_arr, format='PNG')
            img_bytes = img_byte_arr.getvalue()
            
            safe_name = re.sub(r'[^a-zA-Z0-9]', '_', chapter_name).upper()
            file_name = f"{safe_name}.png"
            
            # Step 3: Automatically upload to GCS
            gcs_path = f"module_icons/{file_name}"
            icon_url = upload_to_gcs(img_bytes, GCS_BUCKET_NAME, gcs_path)
                
        except Exception as e:
            st.error(f"🚨 Image Generation/Upload Error for {chapter_name}: {repr(e)}")
            icon_url = "ERROR_GENERATING_ICON"
            
        # Step 4: Stitch Data Together
        file_ch_match = re.search(r'(?:ch(?:apter)?[\s_-]*|^\s*)(\d+)', pdf_file.name, re.IGNORECASE)
        expected_ch_num = str(int(file_ch_match.group(1))) if file_ch_match else None

        for item in chapter_data:
            formatted_row = {
                "SUBJECT": subject_input,
                "MODULE": item.get("Chapter Name", ""),
                "CHAPTER": item.get("Subtopic Name", ""),
                "Icon": icon_url
            }
            master_data.append(formatted_row)
            
        time.sleep(1.5)
        
    progress_bar.empty()
    
    if not master_data:
        st.error("❌ Critical Failure: Could not extract any data to build the Master Excel file.")
    else:
        df = pd.DataFrame(master_data)
        excel_buffer = io.BytesIO()
        
        # Write to Excel and convert string URLs into formal Hyperlink objects
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Master Index')
            worksheet = writer.sheets['Master Index']
            from openpyxl.utils import get_column_letter
            
            icon_col_idx = None
            for i, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 3
                worksheet.column_dimensions[get_column_letter(i + 1)].width = max_len
                if col == "Icon":
                    icon_col_idx = i + 1
            
            # Convert text URLs in the "Icon" column into interactive clickable hyperlinks
            if icon_col_idx:
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=icon_col_idx)
                    if cell.value and str(cell.value).startswith("http"):
                        cell.hyperlink = cell.value
                        cell.style = 'Hyperlink'

        st.success(f"🎉 Complete! Processed {len(df['MODULE'].unique())} chapters.")
        st.dataframe(df)
        
        st.download_button(
            label="📥 Download Master_Index.xlsx",
            data=excel_buffer.getvalue(),
            file_name="Master_Index.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )